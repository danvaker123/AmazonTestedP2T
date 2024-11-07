import json
import logging
import os
import time
import openpyxl
import yaml
from selenium import webdriver
from selenium.common import NoSuchElementException, TimeoutException, ElementClickInterceptedException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select

# Set up logging configuration
logging.basicConfig(filename='automation_log.txt', level=logging.INFO,
                    format='%(asctime)s | %(levelname)s | %(message)s', filemode='w')


# Clear sheet content
def clear_sheet(sheet):
    sheet.delete_rows(1, sheet.max_row)


# Create Excel report from changes data
def create_excel_report(changes_data, output_file, sheet_name):
    try:
        wb = openpyxl.load_workbook(output_file)
    except FileNotFoundError:
        logging.error("Output Excel file '%s' not found.", output_file)
        raise
    except Exception as e:
        logging.error("Error loading output workbook: %s", e)
        raise

    if sheet_name not in wb.sheetnames:
        logging.error("Sheet '%s' not found in the output workbook.", sheet_name)
        raise ValueError(f"Sheet '{sheet_name}' not found in the output workbook.")

    ws = wb[sheet_name]
    clear_sheet(ws)
    header = ["Task", "Action", "Configuration", "Updated_Column1", "Old_Data1", "New_Data1", "Updated_Column2",
              "Old_Data2", "New_Data2", "Updated_Column3", "Old_Data3", "New_Data3", "Updated_Column4", "Old_Data4",
              "New_Data4", "Result", "Comment"]
    ws.append(header)

    for change in changes_data:
        row = [change.get("Step", ""), change.get("Action", ""), change.get("Configuration", ""),
               change.get("Updated_Column1", ""), change.get("Old_Data1", ""), change.get("New_Data1", ""),
               change.get("Updated_Column2", ""), change.get("Old_Data2", ""), change.get("New_Data2", ""),
               change.get("Updated_Column3", ""), change.get("Old_Data3", ""), change.get("New_Data3", ""),
               change.get("Updated_Column4", ""), change.get("Old_Data4", ""), change.get("New_Data4", ""),
               change.get("Result", ""), change.get("Comment", "")]
        ws.append(row)

    # Adjust column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    try:
        wb.save(output_file)
        logging.info("Excel report created at: %s", output_file)
    except Exception as e:
        logging.error("Failed to save Excel report: %s", e)
        raise


# Convert Excel data to JSON format
def excel_to_json(file_path, sheet_name):
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file '{file_path}' not found.")
    except Exception as e:
        raise Exception(f"Error loading workbook: {e}")

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

    sheet = wb[sheet_name]
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {"Subtask ID": row[0], "Task Name": row[1], "Configuration Name": row[2], "Column Name1": row[3],
                    "Value1": row[4], "Column Name2": row[5], "Value2": row[6], "Column Name3": row[7],
                    "Value3": row[8], "Column Name4": row[9], "Value4": row[10]}
        data.append(row_data)

    return data


# Load JSON and YAML files
def load_json(file_path):
    with open(file_path, 'r') as f:
        return json.load(f)


def load_yaml(file_path):
    """Load a YAML file and return its contents."""
    try:
        with open(file_path, 'r') as f:
            return yaml.safe_load(f)
    except FileNotFoundError:
        logging.error("YAML file '%s' not found.", file_path)
        raise
    except yaml.YAMLError as e:
        logging.error("Error parsing YAML file: %s", e)
        raise


# Log handling
def log_task_execution(task_name, subtask_id, status):
    logging.info("Task: %s, Subtask ID: %s, Status: %s", task_name, subtask_id, status)


def log_error(task_name, subtask_id, error):
    logging.error("Task: %s, Subtask ID: %s, Error: %s", task_name, subtask_id, error)


# Browser Automation
def get_by_type(locator_type):
    locator_mapping = {'id': By.ID, 'xpath': By.XPATH, 'name': By.NAME, 'class_name': By.CLASS_NAME,
                       'css_selector': By.CSS_SELECTOR}
    if locator_type is None:
        logging.error("Locator type is None.")
        raise ValueError("Locator type cannot be None.")
    result = locator_mapping.get(locator_type.lower())
    if result is None:
        logging.error("Locator type '%s' not found in mapping.", locator_type)
        raise ValueError(f"Locator type '{locator_type}' not found in mapping.")
    return result


def wait_for_element(driver, locator_type, locator_value):
    """Wait for an element to be present and return it."""
    return WebDriverWait(driver, 10).until(EC.presence_of_element_located((locator_type, locator_value)))


def delete_printer_records(driver):
    # Find all rows in the table
    try:
        check_table_record = driver.find_elements(By.CSS_SELECTOR, '#ServersTable > table.x1o > tbody > tr')
        length = len(check_table_record)
        logging.info(f"Total records found: {length}")
    except NoSuchElementException as e:
        logging.error("Unable to locate table records.", exc_info=True)
        return

    count = 0
    for i in range(1, length - 1):
        try:
            # Attempt to click on the delete link in the 5th column
            clk = WebDriverWait(driver, 120).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, '#ServersTable > table > tbody > tr > td:nth-child(5) > a'))
            ).click()
            logging.info(f"Record {i}: Clicked on the delete link in the 5th column.")
            time.sleep(5)
        except (TimeoutException, ElementClickInterceptedException):
            logging.warning(
                f"Record {i}: Failed to click on the delete link in the 5th column, attempting the 4th column.")
            try:
                # Attempt to click on the delete link in the 4th column if 5th column click fails
                clk = WebDriverWait(driver, 120).until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, '#ServersTable > table > tbody > tr > td:nth-child(4) > a'))
                ).click()
                logging.info(f"Record {i}: Clicked on the delete link in the 4th column.")
                time.sleep(5)
            except (TimeoutException, ElementClickInterceptedException):
                logging.error(f"Record {i}: Failed to click on the delete link in both 4th and 5th columns.",
                              exc_info=True)
                continue

        # Confirm the deletion by clicking "Yes" in the delete confirmation form
        try:
            close_yes = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, '#deleteForm > table > tbody > tr > td > button:nth-child(3)'))
            ).click()
            logging.info(f"Record {i}: Confirmed deletion by clicking 'Yes'.")
            time.sleep(10)
        except (TimeoutException, ElementClickInterceptedException):
            logging.error(f"Record {i}: Failed to confirm deletion by clicking 'Yes'.", exc_info=True)
            continue

        count += 1
        logging.info(f"Deleted record {count}/{length - 1}.")

    logging.info(f"Total records deleted: {count}/{length - 1}.")



def create_context(task_data):
    """
    Create a context dictionary from task data for template replacement.

    :param task_data: Dictionary containing task data from Excel.
    :return: Context dictionary with keys matching column names.
    """
    context = {}
    for key in task_data:
        context[key.lower().replace(" ", "_")] = task_data[key]  # Using lower case and underscores for keys
    return context


subtask_occurrence_tracker = {}


def perform_task(driver, task_config, task_data, username, password, url, subtask_id, first_run):
    # If subtask has been encountered before, start from step 20, otherwise start from step 1
    starting_step = 1 if subtask_occurrence_tracker[subtask_id] == 1 else 20

    try:
        # Navigate to the URL only on the first run of the subtask
        if first_run:
            driver.get(url)
            logging.info("Opened URL: %s", url)
            time.sleep(5)  # Wait for the page to load

        # Execute steps based on the current occurrence of the subtask
        for action in task_config['task']['actions']:
            step_no = action['step_no']

            # Skip steps if they are less than the starting step (e.g., start from step 20 for subsequent executions)
            if step_no < starting_step:
                logging.info(f"Skipping step {step_no} for subtask {subtask_id}. Already executed.")
                continue

            # Perform the action (send_keys, click, etc.)
            perform_action(driver, action, task_data, username, password)

        logging.info("Task '%s' with Subtask '%s' executed successfully.", task_data['Task Name'], task_data['Configuration Name'])


    except Exception as e:
        logging.error("Error occurred during task execution: %s", e)


def perform_action(driver, action, task_data, username=None, password=None, url=None, output_data=None):
    # Initialize variables
    locator_type = action.get('locator_type')  # Get the locator type directly from the action
    output_data = output_data or {}  # Initialize output_data if not provided

    # Create context for templating
    context = create_context(task_data)

    # Handle wait actions without locator values
    if action['action_type'] == 'wait':
        duration = action.get('duration', 10)  # Default to 10 seconds if duration is not specified
        logging.info("Waiting for %d seconds...", duration)
        time.sleep(duration)
        return output_data  # Exit after waiting since no further action is required

    # Get the locator value
    locator_value = action.get('locator_value')

    # Check if locator_value is None for actions that require it
    if action['action_type'] not in ['switch_to_new_tab', 'wait', 'delete_printer_records'] and locator_value is None:
        logging.error("Locator value is None for action: %s", action)
        return output_data  # Skip the action if the locator is None

    # Replace placeholders in locator_value safely
    if locator_value is not None:
        try:
            locator_value = locator_value.format(**context)
        except KeyError as e:
            logging.error("Missing key in context for formatting locator_value: %s", e)
            return output_data  # Skip the action if key is missing

    # Initialize input_value
    input_value = ""

    # Get the input_field structure
    input_field = action.get('input_field', {})

    # Determine if input is needed
    input_required = input_field is not None and input_field.get('type') in ['command_line', 'input_file', 'fixed']

    # Check for input_field type
    if input_required:
        field_name = input_field.get('field_name')
        if input_field.get('type') == "command_line":
            input_value = {"username": username, "password": password, "url": url}.get(field_name, "")
        elif input_field.get('type') == "input_file":
            input_value = task_data.get(field_name, "")
        elif input_field.get('type') == "fixed":
            input_value = input_field.get('value', "")

    # If input_value is required and is still empty, log and skip the action
    if input_required and not input_value:
        logging.warning("Skipping action '%s' because required input value is missing.", action['description'])
        return output_data

    # Wait for the element to be present if action requires it
    element = None
    if action['action_type'] in ['send_keys', 'click', 'find', 'retrieve_value', 'send_keys_enter', 'select_dropdown', 'switch_to_frame', 'toggle_checkbox']:
        locator_type = get_by_type(action['locator_type'])
        try:
            element = wait_for_element(driver, locator_type, locator_value)
        except Exception as e:
            logging.warning("Could not find element with locator '%s': %s", locator_value, e)
            return output_data  # Skip if element cannot be found

    # Perform actions based on action_type
    try:
        if action['action_type'] == 'send_keys':
            if element is not None:
                element.clear()
                element.send_keys(input_value)
                logging.info("Sent keys to element with locator '%s': %s", locator_value, input_value)
            else:
                logging.warning("Element not found for send_keys action with locator '%s'. Action skipped.", locator_value)

        elif action['action_type'] == 'send_keys_enter':
            if element is not None:
                element.send_keys(Keys.ENTER)
                logging.info("Sent ENTER to element with locator '%s'", locator_value)
            else:
                logging.warning("Element not found for send_keys_enter action with locator '%s'. Action skipped.", locator_value)

        elif action['action_type'] == 'click':
            if element is not None:
                element.click()
                logging.info("Clicked element with locator '%s'", locator_value)
            else:
                logging.warning("Element not found for click action with locator '%s'. Action skipped.", locator_value)

        elif action['action_type'] == 'retrieve_value':
            if element is not None:
                retrieved_value = element.get_attribute('value')
                field_name = action.get('output_field', 'retrieved_value')
                output_data[field_name] = retrieved_value
                logging.info("Retrieved value from element with locator '%s': %s", locator_value, retrieved_value)
            else:
                logging.warning("Element not found for retrieve_value action with locator '%s'. Action skipped.", locator_value)

        elif action['action_type'] == 'select_dropdown':
            dropdown_value = action.get('dropdown_value')
            if element is not None and dropdown_value is not None:
                try:
                    select = Select(element)
                    current_selection = select.first_selected_option.text  # Check current selected option
                    if current_selection != dropdown_value:  # Only select if different from current
                        select.select_by_visible_text(dropdown_value)
                        logging.info("Selected '%s' from dropdown with locator '%s'", dropdown_value, locator_value)
                    else:
                        logging.info("Dropdown already set to '%s'; no action taken for locator '%s'", dropdown_value, locator_value)
                except Exception as e:
                    logging.error("Could not select '%s' from dropdown with locator '%s'. Error: %s", dropdown_value, locator_value, str(e))
            else:
                logging.warning("Dropdown element not found or dropdown_value missing for action with locator '%s'. Action skipped.", locator_value)

        elif action['action_type'] == 'switch_to_new_tab':
            original_tab_handle = driver.current_window_handle
            logging.info("Original tab handle: %s", original_tab_handle)
            WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)
            all_handles = driver.window_handles
            new_tab_handle = [handle for handle in all_handles if handle != original_tab_handle]
            if new_tab_handle:
                driver.switch_to.window(new_tab_handle[0])
                logging.info("Switched to the new tab with handle: %s", new_tab_handle[0])
            else:
                logging.warning("New tab not found. Staying on the original tab.")

        elif action['action_type'] == 'switch_to_frame':
            if element is not None:
                driver.switch_to.frame(element)
                logging.info("Switched to iframe with locator '%s'", locator_value)
            else:
                logging.warning("Element not found for switch_to_frame action with locator '%s'. Action skipped.", locator_value)

        elif action['action_type'] == 'delete_printer_records':
            delete_printer_records(driver)  # Call the delete function

        elif action['action_type'] == 'toggle_checkbox':
            desired_state = action.get('desired_state', 'check')  # Default to checking if not specified
            if element is not None:
                # Check the current state of the checkbox using JavaScript
                is_checked = driver.execute_script("return arguments[0].checked;", element)

                if desired_state == 'check' and not is_checked:
                    try:
                        # Scroll the checkbox into view
                        driver.execute_script("arguments[0].scrollIntoView(true);", element)

                        # Click the checkbox if unchecked
                        element.click()
                        logging.info("Checked the checkbox with locator '%s'", locator_value)

                    except Exception as e:
                        logging.error("Error while trying to check checkbox with locator '%s': %s", locator_value, e)

                elif desired_state == 'uncheck' and is_checked:
                    try:
                        # Scroll the checkbox into view
                        driver.execute_script("arguments[0].scrollIntoView(true);", element)

                        # Click the associated label to uncheck the checkbox
                        label_element = driver.find_element(By.XPATH, f"//label[@for='{element.get_attribute('id')}']")
                        label_element.click()
                        logging.info("Unchecked the checkbox using the label for locator '%s'", locator_value)

                    except Exception as e:
                        logging.error("Error while trying to uncheck checkbox with label '%s': %s", locator_value, e)

                elif desired_state == 'uncheck' and not is_checked:
                    logging.info("Checkbox with locator '%s' is already unchecked. No action needed.", locator_value)

                elif desired_state == 'check' and is_checked:
                    logging.info("Checkbox with locator '%s' is already checked. No action needed.", locator_value)

            else:
                logging.warning(
                    "Checkbox element not found for toggle_checkbox action with locator '%s'. Action skipped.",
                    locator_value)

        print(f"Executed step: {action['step_no']} , {action['description']}")

    except Exception as e:
        logging.error("An error occurred while performing action '%s': %s", action['description'], e)

    print(output_data)

    return output_data



def main(input_file_path, output_file_path, sheet_name, config_file_path):
    logging.info("Automation process started.")

    # Load task data from Excel
    task_data = excel_to_json(input_file_path, sheet_name)

    # Load configuration from YAML
    with open(config_file_path, 'r') as f:
        task_config = yaml.safe_load(f)

    username = os.getenv('AUTOMATION_USERNAME', 'casey.brown')  # Load from environment variable
    password = os.getenv('AUTOMATION_PASSWORD', 'Ks?6BK%8')  # Load from environment variable
    url = "https://fa-etan-dev14-saasfademo1.ds-fa.oraclepdemos.com/"

    # Variable to track the current browser and subtask
    current_browser = None
    current_subtask_id = None
    is_first_run = True

    for data in task_data:
        subtask_id = str(data['Subtask ID'])

        # Check if the subtask ID is present in the YAML config
        if subtask_id not in task_config['tasks']:
            logging.error(f"Subtask ID {subtask_id} not found in the configuration.")
            continue  # Skip this subtask if not found in the YAML config

        # Check if the subtask has been processed before
        if subtask_id not in subtask_occurrence_tracker:
            subtask_occurrence_tracker[subtask_id] = 1  # Mark as first occurrence
        else:
            subtask_occurrence_tracker[subtask_id] += 1  # Increment occurrence count

        # If the subtask is different from the last one, close the current browser and open a new one
        if current_subtask_id != subtask_id:
            if current_browser is not None:
                current_browser.quit()
                logging.info(f"Closed browser for subtask {current_subtask_id}. Starting new subtask {subtask_id}.")
            current_browser = webdriver.Chrome()  # Start a new browser session for the new subtask
            current_subtask_id = subtask_id  # Update current subtask ID
            is_first_run = True  # Set to True for a new subtask

        # Perform the task with the current browser
        perform_task(current_browser, task_config['tasks'][subtask_id], data, username, password, url, subtask_id,
                     is_first_run)

        # After performing the task, set is_first_run to False
        is_first_run = False
    # After all subtasks are processed, close the browser
    if current_browser is not None:
        current_browser.quit()
        logging.info("Closed the final browser.")


if __name__ == "__main__":
    main('../Input/input_data.xlsx', '../Input/output_data.xlsx', 'Input Details', '../Config/config1.yaml')
