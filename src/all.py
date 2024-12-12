import json
import logging
import os
import time

import openpyxl
import pandas as pd
import yaml
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.runtime.auth.client_credential import ClientCredential
from office365.sharepoint.client_context import ClientContext
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from selenium import webdriver
from selenium.common import NoSuchElementException, TimeoutException, ElementClickInterceptedException
from selenium.webdriver import Keys
from selenium.webdriver.chrome.service import Service as ChromeService, Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager


# Set up logging configuration
logging.basicConfig(filename='../Logs/automation_log.txt', level=logging.INFO,
                    format='%(asctime)s | %(levelname)s | %(message)s', filemode='w')


# Clear sheet content
def clear_sheet(sheet):
    sheet.delete_rows(1, sheet.max_row)

def functional_log_maker(input_file_path):
        file_path = input_file_path

        # Load Excel data to get task and configuration names
        df = pd.read_excel(file_path)
        task_names = df['Task Name'].to_list()
        config_names = df['Configuration Name'].to_list()

        # Define the log file path
        log_file_path = "../Logs/automation_log.txt"

        # Read the content of the log file
        with open(log_file_path, 'r') as file:
            content = file.readlines()

        # Create a list to store structured logs
        function_log = []
        logged_warnings = set()

        # Iterate over task and configuration pairs to process each section
        for idx, task in enumerate(task_names):
            # Get the configuration, and handle NaN or empty values gracefully
            config = config_names[idx] if idx < len(config_names) else ""
            if pd.isna(config) or config == "":
                config = "None"

            section_header = f"\n--- Task: {task} | Configuration: {config} ---\n"
            function_log.append(section_header)

            # Flags for tracking task state
            task_error_found = False
            task_success_found = False

            # Iterate over the log content to process warnings, errors, and success messages
            for line in content:
                # Capture warnings relevant to the task
                if "WARNING" in line and "Could not find element with locator" not in line:
                    if line.strip() not in logged_warnings:
                        function_log.append("    WARNING: " + line.strip() + "\n")
                        logged_warnings.add(line.strip())

                # Capture any error for the task
                if "ERROR" in line and (
                        f"Task '{task}'" in line or (config != "None" and f"Configuration '{config}'" in line)):
                    function_log.append("    ERROR: " + line.strip() + "\n")
                    task_error_found = True  # Mark an error was found

                # Capture success message only if no prior errors for the task
                if f"Task '{task}' with Subtask '{config}' executed successfully" in line:
                    if not task_error_found:
                        function_log.append("    SUCCESS: " + line.strip() + "\n")
                        task_success_found = True
                    break  # Stop further checking after a successful execution is logged

            # Finalize task status in the functional log
            if not task_success_found and not task_error_found:
                function_log.append(f"    ERROR: Task '{task}' did not execute successfully.\n")

        # Write the structured log to a new file
        output_file = "../Logs/functional_log.txt"
        with open(output_file, 'w') as file:
            for entry in function_log:
                file.write(entry)

# Create Excel report from changes data
def create_excel_report(changes_data, output_file, sheet_name):
    try:
        wb = openpyxl.load_workbook(output_file)
    except FileNotFoundError:
        logging.error("Output Excel file '%s' not found.", output_file)
        raise
    except Exception as e:
        logging.error("Error loading output workbook: %s", e)
        raise

    if sheet_name not in wb.sheetnames:
        logging.error("Sheet '%s' not found in the output workbook.", sheet_name)
        raise ValueError(f"Sheet '{sheet_name}' not found in the output workbook.")

    ws = wb[sheet_name]
    clear_sheet(ws)
    header = ["Task", "Action", "Configuration", "Updated_Column1", "Old_Data1", "New_Data1", "Updated_Column2",
              "Old_Data2", "New_Data2", "Updated_Column3", "Old_Data3", "New_Data3", "Updated_Column4", "Old_Data4",
              "New_Data4", "Result", "Comment"]
    ws.append(header)

    for change in changes_data:
        row = [change.get("Step", ""), change.get("Action", ""), change.get("Configuration", ""),
               change.get("Updated_Column1", ""), change.get("Old_Data1", ""), change.get("New_Data1", ""),
               change.get("Updated_Column2", ""), change.get("Old_Data2", ""), change.get("New_Data2", ""),
               change.get("Updated_Column3", ""), change.get("Old_Data3", ""), change.get("New_Data3", ""),
               change.get("Updated_Column4", ""), change.get("Old_Data4", ""), change.get("New_Data4", ""),
               change.get("Result", ""), change.get("Comment", "")]
        ws.append(row)

    # Adjust column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    try:
        wb.save(output_file)
        logging.info("Excel report created at: %s", output_file)
    except Exception as e:
        logging.error("Failed to save Excel report: %s", e)
        raise


# Convert Excel data to JSON format
def excel_to_json(file_path, sheet_name):
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file '{file_path}' not found.")
    except Exception as e:
        raise Exception(f"Error loading workbook: {e}")

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

    sheet = wb[sheet_name]
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {"Subtask ID": row[0], "Task Name": row[1], "Configuration Name": row[2], "Column Name1": row[3],
                    "Value1": row[4], "Column Name2": row[5], "Value2": row[6], "Column Name3": row[7],
                    "Value3": row[8], "Column Name4": row[9], "Value4": row[10]}
        data.append(row_data)

    return data


# Load JSON and YAML files
def load_json(file_path):
    with open(file_path, 'r') as f:
        return json.load(f)


def load_yaml(file_path):
    """Load a YAML file and return its contents."""
    try:
        with open(file_path, 'r') as f:
            return yaml.safe_load(f)
    except FileNotFoundError:
        logging.error("YAML file '%s' not found.", file_path)
        raise
    except yaml.YAMLError as e:
        logging.error("Error parsing YAML file: %s", e)
        raise


# Log handling
def log_task_execution(task_name, subtask_id, status):
    logging.info("Task: %s, Subtask ID: %s, Status: %s", task_name, subtask_id, status)


def log_error(task_name, subtask_id, error):
    logging.error("Task: %s, Subtask ID: %s, Error: %s", task_name, subtask_id, error)


# Browser Automation
def get_by_type(locator_type):
    locator_mapping = {'id': By.ID, 'xpath': By.XPATH, 'name': By.NAME, 'class_name': By.CLASS_NAME,
                       'css_selector': By.CSS_SELECTOR}
    if locator_type is None:
        logging.error("Locator type is None.")
        raise ValueError("Locator type cannot be None.")
    result = locator_mapping.get(locator_type.lower())
    if result is None:
        logging.error("Locator type '%s' not found in mapping.", locator_type)
        raise ValueError(f"Locator type '{locator_type}' not found in mapping.")
    return result


def wait_for_element(driver, locator_type, locator_value):
    """Wait for an element to be present and return it."""
    return WebDriverWait(driver, 10).until(EC.presence_of_element_located((locator_type, locator_value)))


def clear_text_box(driver, locator_type, locator_value):
    try:
        # Locate the element using the provided locator_type and locator_value
        element = wait_for_element(driver, locator_type, locator_value)

        if element is not None:
            # Clear the text box
            element.clear()
            logging.info("Cleared text box with locator '%s'", locator_value)
        else:
            logging.warning("Element not found for clearing text box with locator '%s'. Action skipped.", locator_value)

    except Exception as e:
        logging.error("An error occurred while clearing text box with locator '%s': %s", locator_value, e)


def delete_printer_records(driver):
    # Find all rows in the table
    try:
        check_table_record = driver.find_elements(By.CSS_SELECTOR, '#ServersTable > table.x1o > tbody > tr')
        length = len(check_table_record)
        logging.info(f"Total records found: {length}")
    except NoSuchElementException as e:
        logging.error("Unable to locate table records.", exc_info=True)
        return

    count = 0
    for i in range(1, length - 1):
        try:
            # Attempt to click on the delete link in the 5th column
            clk = WebDriverWait(driver, 120).until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, '#ServersTable > table > tbody > tr > td:nth-child(5) > a'))).click()
            logging.info(f"Record {i}: Clicked on the delete link in the 5th column.")
            time.sleep(5)
        except (TimeoutException, ElementClickInterceptedException):
            logging.warning(
                f"Record {i}: Failed to click on the delete link in the 5th column, attempting the 4th column.")
            try:
                # Attempt to click on the delete link in the 4th column if 5th column click fails
                clk = WebDriverWait(driver, 120).until(EC.presence_of_element_located(
                    (By.CSS_SELECTOR, '#ServersTable > table > tbody > tr > td:nth-child(4) > a'))).click()
                logging.info(f"Record {i}: Clicked on the delete link in the 4th column.")
                time.sleep(5)
            except (TimeoutException, ElementClickInterceptedException):
                logging.error(f"Record {i}: Failed to click on the delete link in both 4th and 5th columns.",
                              exc_info=True)
                continue

        # Confirm the deletion by clicking "Yes" in the delete confirmation form
        try:
            close_yes = WebDriverWait(driver, 60).until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, '#deleteForm > table > tbody > tr > td > button:nth-child(3)'))).click()
            logging.info(f"Record {i}: Confirmed deletion by clicking 'Yes'.")
            time.sleep(10)
        except (TimeoutException, ElementClickInterceptedException):
            logging.error(f"Record {i}: Failed to confirm deletion by clicking 'Yes'.", exc_info=True)
            continue

        count += 1
        logging.info(f"Deleted record {count}/{length - 1}.")

    logging.info(f"Total records deleted: {count}/{length - 1}.")


def create_context(task_data):
    """
    Create a context dictionary from task data for template replacement.

    :param task_data: Dictionary containing task data from Excel.
    :return: Context dictionary with keys matching column names.
    """
    context = {}
    for key in task_data:
        context[key.lower().replace(" ", "_")] = task_data[key]  # Using lower case and underscores for keys
    return context


subtask_occurrence_tracker = {}


def perform_task(driver, task_config, task_data, username, password, url, subtask_id, first_run):
    # If subtask has been encountered before, start from step 20, otherwise start from step 1
    starting_step = 1 if subtask_occurrence_tracker[subtask_id] == 1 else 20
    output_data = {}

    try:
        if first_run:
            driver.get(url)
            logging.info("Opened URL: %s", url)
            time.sleep(5)

        for action in task_config['task']['actions']:
            step_no = action['step_no']

            if step_no < starting_step:
                logging.info(f"Skipping step {step_no} for subtask {subtask_id}. Already executed.")
                continue

            result = perform_action(driver, action, task_data, username, password, output_data)

            if result and action['action_type'] == 'retrieve_value':
                output_data.update(result)

        logging.info("Task '%s' with Subtask '%s' executed successfully.", task_data['Task Name'],
                     task_data['Configuration Name'])

    except Exception as e:
        logging.error("Error occurred during task execution: %s", e)

    return output_data


def perform_action(driver, action, task_data, username=None, password=None, url=None, output_data=None):
    # Initialize variables
    locator_type = action.get('locator_type')  # Get the locator type directly from the action
    output_data = output_data or {}  # Initialize output_data if not provided

    # Create context for templating
    context = create_context(task_data)

    # Handle wait actions without locator values
    if action['action_type'] == 'wait':
        duration = action.get('duration', 10)  # Default to 10 seconds if duration is not specified
        logging.info("Waiting for %d seconds...", duration)
        time.sleep(duration)
        return output_data  # Exit after waiting since no further action is required

    # Get the locator value
    locator_value = action.get('locator_value')

    # Check if locator_value is None for actions that require it
    if action['action_type'] not in ['switch_to_new_tab', 'wait', 'delete_printer_records'] and locator_value is None:
        logging.error("Locator value is None for action: %s", action)
        return output_data  # Skip the action if the locator is None

    # Replace placeholders in locator_value safely
    if locator_value is not None:
        try:
            locator_value = locator_value.format(**context)
        except KeyError as e:
            logging.error("Missing key in context for formatting locator_value: %s", e)
            return output_data  # Skip the action if key is missing

    # Initialize input_value
    input_value = ""

    # Get the input_field structure
    input_field = action.get('input_field', {})

    # Determine if input is needed
    input_required = input_field is not None and input_field.get('type') in ['command_line', 'input_file', 'fixed']

    # Check for input_field type
    if input_required:
        field_name = input_field.get('field_name')
        if input_field.get('type') == "command_line":
            input_value = {"username": username, "password": password, "url": url}.get(field_name, "")
        elif input_field.get('type') == "input_file":
            input_value = task_data.get(field_name, "")
        elif input_field.get('type') == "fixed":
            input_value = input_field.get('value', "")

    # If input_value is required and is still empty, log and skip the action
    if input_required and not input_value:
        logging.warning("Skipping action '%s' because required input value is missing.", action['description'])
        return output_data

    # Wait for the element to be present if action requires it
    element = None
    if action['action_type'] in ['send_keys', 'click', 'find', 'retrieve_value', 'send_keys_enter', 'select_dropdown',
                                 'switch_to_frame', 'toggle_checkbox', 'clear_text_box']:
        locator_type = get_by_type(action['locator_type'])
        try:
            element = wait_for_element(driver, locator_type, locator_value)
        except Exception as e:
            logging.warning("Could not find element with locator '%s': %s", locator_value, e)
            return output_data  # Skip if element cannot be found

    # Perform actions based on action_type
    try:
        if action['action_type'] == 'send_keys':
            if element is not None:
                element.clear()
                element.send_keys(input_value)
                logging.info("Sent keys to element with locator '%s': %s", locator_value, input_value)
            else:
                logging.warning("Element not found for send_keys action with locator '%s'. Action skipped.",
                                locator_value)

        elif action['action_type'] == 'send_keys_enter':
            if element is not None:
                element.send_keys(Keys.ENTER)
                logging.info("Sent ENTER to element with locator '%s'", locator_value)
            else:
                logging.warning("Element not found for send_keys_enter action with locator '%s'. Action skipped.",
                                locator_value)

        elif action['action_type'] == 'click':
            if element is not None:
                element.click()
                logging.info("Clicked element with locator '%s'", locator_value)
            else:
                logging.warning("Element not found for click action with locator '%s'. Action skipped.", locator_value)

        if action['action_type'] == 'retrieve_value':
            if element is not None:
                retrieved_value = element.get_attribute('value')
                field_name = action.get('output_field', 'old_value_column1')  # Set dynamically, e.g., old_value_column2
                output_data[field_name] = retrieved_value
                logging.info("Retrieved value from element with locator '%s': %s", locator_value, retrieved_value)
            else:
                logging.warning("Element not found for retrieve_value action with locator '%s'. Action skipped.",
                                locator_value)

        elif action['action_type'] == 'select_dropdown':
            dropdown_value = action.get('dropdown_value')
            if element is not None and dropdown_value is not None:
                try:
                    select = Select(element)
                    current_selection = select.first_selected_option.text  # Check current selected option
                    if current_selection != dropdown_value:  # Only select if different from current
                        select.select_by_visible_text(dropdown_value)
                        logging.info("Selected '%s' from dropdown with locator '%s'", dropdown_value, locator_value)
                    else:
                        logging.info("Dropdown already set to '%s'; no action taken for locator '%s'", dropdown_value,
                                     locator_value)
                except Exception as e:
                    logging.error("Could not select '%s' from dropdown with locator '%s'. Error: %s", dropdown_value,
                                  locator_value, str(e))
            else:
                logging.warning(
                    "Dropdown element not found or dropdown_value missing for action with locator '%s'. Action skipped.",
                    locator_value)

        elif action['action_type'] == 'switch_to_new_tab':
            original_tab_handle = driver.current_window_handle
            logging.info("Original tab handle: %s", original_tab_handle)
            WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)
            all_handles = driver.window_handles
            new_tab_handle = [handle for handle in all_handles if handle != original_tab_handle]
            if new_tab_handle:
                driver.switch_to.window(new_tab_handle[0])
                logging.info("Switched to the new tab with handle: %s", new_tab_handle[0])
            else:
                logging.warning("New tab not found. Staying on the original tab.")

        elif action['action_type'] == 'switch_to_frame':
            if element is not None:
                driver.switch_to.frame(element)
                logging.info("Switched to iframe with locator '%s'", locator_value)
            else:
                logging.warning("Element not found for switch_to_frame action with locator '%s'. Action skipped.",
                                locator_value)

        elif action['action_type'] == 'delete_printer_records':
            delete_printer_records(driver)  # Call the delete function

        elif action['action_type'] == 'toggle_checkbox':
            desired_state = action.get('desired_state', 'check')  # Default to checking if not specified
            if element is not None:
                # Check the current state of the checkbox using JavaScript
                is_checked = driver.execute_script("return arguments[0].checked;", element)

                if desired_state == 'check' and not is_checked:
                    try:
                        # Scroll the checkbox into view
                        driver.execute_script("arguments[0].scrollIntoView(true);", element)

                        # Click the checkbox if unchecked
                        element.click()
                        logging.info("Checked the checkbox with locator '%s'", locator_value)

                    except Exception as e:
                        logging.error("Error while trying to check checkbox with locator '%s': %s", locator_value, e)

                elif desired_state == 'uncheck' and is_checked:
                    try:
                        # Scroll the checkbox into view
                        driver.execute_script("arguments[0].scrollIntoView(true);", element)

                        # Click the associated label to uncheck the checkbox
                        label_element = driver.find_element(By.XPATH, f"//label[@for='{element.get_attribute('id')}']")
                        label_element.click()
                        logging.info("Unchecked the checkbox using the label for locator '%s'", locator_value)

                    except Exception as e:
                        logging.error("Error while trying to uncheck checkbox with label '%s': %s", locator_value, e)

                elif desired_state == 'uncheck' and not is_checked:
                    logging.info("Checkbox with locator '%s' is already unchecked. No action needed.", locator_value)

                elif desired_state == 'check' and is_checked:
                    logging.info("Checkbox with locator '%s' is already checked. No action needed.", locator_value)

            else:
                logging.warning(
                    "Checkbox element not found for toggle_checkbox action with locator '%s'. Action skipped.",
                    locator_value)

        elif action['action_type'] == 'clear_text_box':
            if element is not None:
                element.clear()
                logging.info("Cleared text box with locator '%s'", locator_value)
            else:
                logging.warning("Element not found for clear_text_box action with locator '%s'. Action skipped.",
                                locator_value)

        print(f"Executed step: {action['step_no']} , {action['description']}")

    except Exception as e:
        logging.error("An error occurred while performing action '%s': %s", action['description'], e)

    print(output_data)

    return output_data


def download_file_from_sharepoint(site_url, client_id, client_secret, sharepoint_file_path, local_file_path):
    ctx = ClientContext(site_url).with_credentials(ClientCredential(client_id, client_secret))
    file = ctx.web.get_file_by_server_relative_url(sharepoint_file_path)
    with open(local_file_path, 'wb') as local_file:
        file.download(local_file).execute_query()
    print(f"Downloaded {sharepoint_file_path} to {local_file_path}")

def upload_file_to_sharepoint(site_url, client_id, client_secret, sharepoint_file_path, local_file_path):
    ctx = ClientContext(site_url).with_credentials(ClientCredential(client_id, client_secret))
    with open(local_file_path, 'rb') as local_file:
        file_content = local_file.read()
    folder_path = f"/sites/{site_url.split('/')[-1]}/Shared Documents/Testing"
    file_name = os.path.basename(local_file_path)
    target_folder = ctx.web.get_folder_by_server_relative_url(folder_path)
    target_file = target_folder.upload_file(file_name, file_content)
    ctx.execute_query()
    print(f"Uploaded {local_file_path} to {folder_path}")


def start_browser():
    chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_argument("--headless")  # Enable headless mode
    chrome_options.add_argument("--no-sandbox")  # Bypass sandboxing
    chrome_options.add_argument("--disable-dev-shm-usage")  # Reduce memory usage
    chrome_options.add_argument("window-size=800,600")  # Smaller window size
    chrome_options.add_argument('--disable-gpu')  # Disable GPU acceleration
    chrome_options.add_argument('--disable-browser-side-navigation')  # Disable browser-side navigation

    logging.info("Starting browser in headless mode...")
    try:
        # Use the latest ChromeDriver
        driver = webdriver.Remote(command_executor='http://localhost:4444/wd/hub',
                                  options=chrome_options)

        # Increase the implicit wait timeout to give the browser more time to respond
        driver.implicitly_wait(180)  # Increased timeout

        logging.info("Browser launched successfully!")
        return driver
    except Exception as e:
        logging.error("Error launching browser: %s", e)
        return None


def main(input_file_path, output_file_path, sheet_name, config_file_path, username, password, url):
    logging.info("Automation process started.")


    # Load task data from Excel
    task_data = excel_to_json(input_file_path, sheet_name)
    df_output = pd.read_excel(output_file_path)

    # Load configuration from YAML
    with open(config_file_path, 'r') as f:
        task_config = yaml.safe_load(f)

    # Variable to track the current browser and subtask
    current_browser = None
    current_subtask_id = None
    is_first_run = True
    output_list = []

    # Define the column order you want in the output Excel file
    column_order = ['Task', 'Action', 'Configuration', 'Updated_Column1', 'New_Data1', 'Old_Data1', 'Updated_Column2',
                    'New_Data2', 'Old_Data2', 'Updated_Column3', 'New_Data3', 'Old_Data3', 'Updated_Column4',
                    'New_Data4', 'Old_Data4', 'Result', 'Comment']

    for data in task_data:
        subtask_id = str(data['Subtask ID'])

        # Ensure output_dict is freshly initialized for each task
        output_dict = {"Task": str(data['Task Name']), "Action": 'Updating Configuration ' + str(
            data['Configuration Name']) + ' Updating Column ' + str(data['Column Name1']) + ' , ' + str(
            data['Column Name2']) + ' , ' + str(data['Column Name3']) + ' , ' + str(data['Column Name4']),
                       'Configuration': str(data['Configuration Name']), 'Updated_Column1': str(data['Column Name1']),
                       'New_Data1': str(data['Value1']), 'Updated_Column2': str(data['Column Name2']),
                       'New_Data2': str(data['Value2']), 'Updated_Column3': str(data['Column Name3']),
                       'New_Data3': str(data['Value3']), 'Updated_Column4': str(data['Column Name4']),
                       'New_Data4': str(data['Value4']), 'Old_Data1': '',  # Initialize Old_Data columns as empty
                       'Old_Data2': '', 'Old_Data3': '', 'Old_Data4': '', 'Result': '', 'Comment': ''}

        # Check if the subtask ID is present in the YAML config
        if subtask_id not in task_config['tasks']:
            logging.error(f"Subtask ID {subtask_id} not found in the configuration.")
            continue  # Skip this subtask if not found in the YAML config

        # Check if the subtask has been processed before
        if subtask_id not in subtask_occurrence_tracker:
            subtask_occurrence_tracker[subtask_id] = 1  # Mark as first occurrence
        else:
            subtask_occurrence_tracker[subtask_id] += 1  # Increment occurrence count

        # If the subtask is different from the last one, close the current browser and open a new one
        if current_subtask_id != subtask_id:
            if current_browser is not None:
                current_browser.quit()
                logging.info(f"Closed browser for subtask {current_subtask_id}. Starting new subtask {subtask_id}.")
            current_browser = start_browser() # Start a new browser session for the new subtask
            current_subtask_id = subtask_id  # Update current subtask ID
            is_first_run = True  # Set to True for a new subtask

        # Perform the task with the current browser
        task_output_data = perform_task(current_browser, task_config['tasks'][subtask_id], data, username, password,
                                        url, subtask_id, is_first_run)

        # After performing the task, set is_first_run to False
        is_first_run = False

        # Update the output_dict with any retrieved values from perform_task (like Old_Data1, Old_Data2)
        for i in range(1, 5):
            field_name = f'old_value_column{i}'
            if field_name in task_output_data:
                output_dict[f'Old_Data{i}'] = task_output_data[field_name]

        # Append the updated output_dict to the output_list for this subtask
        output_list.append(output_dict)

    # After all subtasks are processed, close the browser
    if current_browser is not None:
        current_browser.quit()
        logging.info("Closed the final browser.")

    # Convert the output list to a DataFrame and ensure the correct column order
    output_data = pd.DataFrame(output_list)

    # Reorder the columns according to the specified order
    output_data = output_data[column_order]

    # Save the final output to Excel
    output_data.to_excel(output_file_path, index=False)

    # Now apply formatting with openpyxl for better readability
    wb = openpyxl.load_workbook(output_file_path)
    ws = wb.active

    # Apply bold font to the header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Apply borders to all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Apply alternating row colors (light gray)
    for row in range(2, ws.max_row + 1):
        if row % 2 == 0:
            for cell in ws[row]:
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Adding some padding for the content
        ws.column_dimensions[column].width = adjusted_width

    # Save the formatted Excel file
    wb.save(output_file_path)
    functional_log_maker(input_file_path)


if __name__ == "__main__":
    main('../Input/input_data.xlsx', '../Input/output_data.xlsx', 'Input Details', '../Config/config1.yaml',
         'Casey.Brown', 'u^5X#rP6',
         'https://fa-etan-dev14-saasfademo1.ds-fa.oraclepdemos.com')  # parser = argparse.ArgumentParser(description="Automation Script Arguments")  # parser.add_argument("--input_file", required=True, help="Path to the input Excel file.")  # parser.add_argument("--sheet_name", required=True, help="Sheet name in the Excel input file.")  # parser.add_argument("--username", required=True, help="Automation username.")  # parser.add_argument("--password", required=True, help="Automation password.")  # parser.add_argument("--url", required=True, help="URL to be automated.")  #  # args = parser.parse_args()  # config_file_path = '../Config/config1.yaml'  # output_file_path = '../Input/output_data.xlsx'  #  #  #  # main(  #     input_file_path=args.input_file,  #     output_file_path= output_file_path,  #     sheet_name=args.sheet_name,  #     config_file_path=config_file_path,  #     username=args.username,  #     password=args.password,  #     url=args.url,  # )
